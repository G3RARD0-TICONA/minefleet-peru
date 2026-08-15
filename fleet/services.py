from datetime import date, timedelta
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook

from core.models import AuditLog
from .models import Vehicle, VehicleDocument


VEHICLE_HEADERS = [
    "placa", "vin", "motor", "marca", "modelo", "año", "tipo", "modalidad",
    "propietario", "tara_kg", "pbv_kg", "kilometraje", "estado",
]


def log_action(*, company, user, action, entity, object_id="", detail=None):
    AuditLog.objects.create(
        company=company, user=user, action=action, entity=entity,
        object_id=str(object_id), detail=detail or {},
    )


@transaction.atomic
def import_vehicles(*, workbook_file, company, user):
    workbook = load_workbook(workbook_file, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
    missing = [header for header in VEHICLE_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")
    positions = {name: headers.index(name) for name in VEHICLE_HEADERS}
    created, updated, errors = 0, 0, []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            data = {name: row[index] for name, index in positions.items()}
            defaults = {
                "vin": str(data["vin"]).strip(), "engine_number": str(data["motor"] or "").strip(),
                "brand": str(data["marca"]).strip(), "model": str(data["modelo"]).strip(),
                "year": int(data["año"]), "vehicle_type": str(data["tipo"] or "Volquete").strip(),
                "availability": str(data["modalidad"] or Vehicle.Availability.OWNED).strip().upper(),
                "owner_name": str(data["propietario"]).strip(), "tare_kg": data["tara_kg"],
                "gross_weight_kg": data["pbv_kg"], "odometer_km": int(data["kilometraje"] or 0),
                "status": str(data["estado"] or Vehicle.Status.DRAFT).strip().upper(),
            }
            vehicle, was_created = Vehicle.objects.update_or_create(
                company=company, plate=str(data["placa"]).strip().upper(), defaults=defaults,
            )
            created += int(was_created)
            updated += int(not was_created)
            log_action(company=company, user=user, action="IMPORT_CREATE" if was_created else "IMPORT_UPDATE", entity="Vehicle", object_id=vehicle.pk)
        except Exception as exc:
            errors.append(f"Fila {row_number}: {exc}")
    return {"created": created, "updated": updated, "errors": errors}


def vehicle_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vehiculos"
    sheet.append(VEHICLE_HEADERS)
    sheet.append(["V0X-001", "1MNFLEETDEMO0001", "MTR-DEMO-01", "Volvo", "FMX", 2023, "Volquete", "OWNED", "Transportes Demo SAC", 15000, 41000, 25000, "DRAFT"])
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True, color="FFFFFF")
        cell.fill = cell.fill.copy(fill_type="solid", fgColor="17324D")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = max(12, min(28, max(len(str(c.value or "")) for c in column) + 2))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def refresh_document_alerts(company):
    changed = 0
    for document in VehicleDocument.objects.filter(vehicle__company=company):
        new_status = document.refresh_status()
        if new_status != document.status:
            VehicleDocument.objects.filter(pk=document.pk).update(status=new_status)
            changed += 1
    return changed


def dashboard_metrics(company):
    refresh_document_alerts(company)
    today = date.today()
    vehicles = Vehicle.objects.filter(company=company)
    documents = VehicleDocument.objects.filter(vehicle__company=company)
    return {
        "vehicles": vehicles.count(),
        "eligible": vehicles.filter(status=Vehicle.Status.ELIGIBLE).count(),
        "blocked": vehicles.filter(status__in=(Vehicle.Status.BLOCKED, Vehicle.Status.MAINTENANCE)).count(),
        "expiring": documents.filter(expiry_date__range=(today, today + timedelta(days=30))).count(),
        "expired": documents.filter(expiry_date__lt=today).count(),
    }
